"""Lectura y escritura de archivos Excel con look premium."""
from io import BytesIO
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Timezone Ecuador (UTC-5) — sin DST
try:
    from zoneinfo import ZoneInfo
    _TZ_EC = ZoneInfo("America/Guayaquil")
except ImportError:
    from datetime import timezone, timedelta
    _TZ_EC = timezone(timedelta(hours=-5))


# ── Paleta de colores premium (branding RUBASA Facility Services) ────────────
# Colores oficiales tomados del logo + sitio web rubasa.com.ec
NAVY        = "0F2C5C"   # Azul navy oscuro — texto "RUBASA" del logo
NAVY_LIGHT  = "1E5BFA"   # Azul vibrante — ícono "R" del logo
ACCENT      = "3B7BFF"   # Azul claro — tono medio para acentos

VERDE_BG    = "D4EFDF"  # APTO
AMARILLO_BG = "FCF3CF"  # OBSERVACIÓN
ROJO_BG     = "F1948A"  # RECHAZAR
CRITICO_BG  = "922B21"  # CRÍTICO (rojo intenso oscuro)
GRIS_BG     = "EAEDED"  # SIN DATOS

ROW_ALT_BG  = "FAFBFC"
HEADER_BG   = NAVY
BANNER_BG   = NAVY

TEXTO_OSCURO = "1C2833"
TEXTO_GRIS   = "7B7D7D"
BORDE_SUTIL  = "D5D8DC"

# ── Estilos reutilizables ─────────────────────────────────────────────────────
FONT_TITLE   = Font(name="Calibri", size=18, bold=True, color="FFFFFF")
FONT_SUB     = Font(name="Calibri", size=10, color="D5D8DC", italic=True)
FONT_HEADER  = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
FONT_STAT    = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
FONT_BODY    = Font(name="Calibri", size=10, color=TEXTO_OSCURO)
FONT_BODY_B  = Font(name="Segoe UI Emoji", size=10, bold=True, color=TEXTO_OSCURO)
FONT_CEDULA  = Font(name="Consolas", size=10, color=TEXTO_OSCURO)

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT   = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)

_BORDER = Border(
    left=Side(style="thin", color=BORDE_SUTIL),
    right=Side(style="thin", color=BORDE_SUTIL),
    top=Side(style="thin", color=BORDE_SUTIL),
    bottom=Side(style="thin", color=BORDE_SUTIL),
)


def leer_cedulas(file_bytes: bytes) -> tuple[list[dict], list[str]]:
    """
    Lee un Excel y devuelve (lista de dicts {cedula, nombre}, errores).

    Detección FLEXIBLE de columnas — pensado para que el usuario pueda
    subir cualquier hoja Excel "improvisada" sin tener que renombrar headers:

      • Cédula  → header contiene "cedula", "cédula", "identif" o es == "ci"
      • Nombre  → header contiene "nombre" o "apellido"
                 (cubre "Nombres", "Apellidos y Nombres", "NombreCompleto",
                  "Nombre Apellido", "NOMBRES Y APELLIDOS", etc.)

    Si no hay header reconocible, asume cédula = columna A y nombre = columna B.

    Además dedupea cédulas repetidas: si una misma cédula aparece varias
    veces, conserva la PRIMERA con nombre no-vacío (o la primera, si todas
    vienen vacías).
    """
    try:
        wb = load_workbook(BytesIO(file_bytes), data_only=True)
        ws = wb.active

        # Buscar columnas en el header (matching por substring)
        col_cedula = None
        col_nombre = None
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
        for i, val in enumerate(header_row):
            if not val:
                continue
            v = str(val).strip().lower()
            # Cédula: substring match para tolerar "N° identificación", "CI", etc.
            if col_cedula is None and (
                "cedula" in v or "cédula" in v or "identif" in v or v == "ci"
            ):
                col_cedula = i
            # Nombre: substring "nombre" o "apellido" — captura todas las
            # variantes posibles que un usuario pondría intuitivamente
            elif col_nombre is None and ("nombre" in v or "apellido" in v):
                col_nombre = i

        if col_cedula is None:
            # Sin header reconocible: asumir cédula en A, nombre en B si existe
            col_cedula = 0
            if col_nombre is None and len(header_row) > 1:
                col_nombre = 1
            data_start_row = 1
        else:
            data_start_row = 2

        # Recolectar + dedupear: la primera aparición con nombre gana
        seen: dict[str, dict] = {}
        orden: list[str] = []
        for row in ws.iter_rows(min_row=data_start_row, values_only=True):
            if col_cedula >= len(row):
                continue
            val = row[col_cedula]
            if val is None or str(val).strip() == "":
                continue
            cedula = str(val).strip()
            if cedula.isdigit() and len(cedula) < 10:
                cedula = cedula.zfill(10)

            nombre = ""
            if col_nombre is not None and col_nombre < len(row):
                n = row[col_nombre]
                if n:
                    nombre = str(n).strip()

            if cedula not in seen:
                seen[cedula] = {"cedula": cedula, "nombre": nombre}
                orden.append(cedula)
            else:
                # Si la entrada existente está vacía de nombre pero esta nueva trae uno, upgrade
                if not seen[cedula]["nombre"] and nombre:
                    seen[cedula]["nombre"] = nombre

        items = [seen[c] for c in orden]
        return items, []

    except Exception as e:
        return [], [f"Error leyendo Excel: {type(e).__name__}: {str(e)[:200]}"]


def _aplicar_borde(ws, rango_min_row, rango_max_row, rango_min_col, rango_max_col):
    for row in ws.iter_rows(min_row=rango_min_row, max_row=rango_max_row,
                            min_col=rango_min_col, max_col=rango_max_col):
        for cell in row:
            cell.border = _BORDER


def generar_excel_resultados(resultados: list[dict], tipo: str, incluir_setec: bool = False,
                             incluir_fiscalia: bool = False) -> bytes:
    """
    Genera un Excel premium con los resultados.
    Estructura: banner + stats + tabla.
    incluir_setec    → agrega 3 columnas al final (Estado SETEC, # Cursos, Cursos)
    incluir_fiscalia → agrega 3 columnas al final (Estado Fiscalía, # Noticias, Delitos)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Resultados"

    # ── Definir columnas según tipo ──────────────────────────────────────────
    if tipo == "bachiller":
        headers = ["Cédula", "Nombre", "Estado", "Título", "Especialidad", "Institución", "Año", "Detalle"]
        widths  = [14, 32, 16, 30, 24, 36, 10, 40]
    elif tipo == "satje":
        headers = ["Cédula", "Nombre", "Estado", "T. Demandado", "T. Actor", "Delitos / Detalle"]
        widths  = [14, 32, 18, 14, 14, 60]
    elif tipo == "setec":
        headers = ["Cédula", "Nombre", "🎖️ Estado SETEC", "# Cursos", "Capacitaciones"]
        widths  = [14, 32, 22, 12, 70]
    else:  # completo
        headers = [
            "Cédula", "Nombre", "🚦 Semáforo",
            "🎓 Bachiller", "Institución", "Año",
            "⚖️ Estado SATJE", "Demandado", "Actor", "Delitos / Detalle",
        ]
        widths  = [14, 30, 14, 26, 32, 10, 18, 12, 12, 50]

    # ── Agregar columnas SETEC al final si fue solicitado y no es tipo "setec" ──
    if incluir_setec and tipo != "setec":
        headers += ["🎖️ Estado SETEC", "# Cursos", "Capacitaciones"]
        widths  += [22, 12, 70]

    # ── Agregar columnas Fiscalía al final si fue solicitado ─────────────────
    if incluir_fiscalia:
        headers += ["🚨 Estado Fiscalía", "# Noticias", "Delitos SIAF"]
        widths  += [22, 12, 50]

    num_cols = len(headers)

    # ── Fila 1: Banner con título y branding ─────────────────────────────────
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    cell = ws.cell(row=1, column=1, value="⚙️ JR Verifica EC")
    cell.fill      = PatternFill(start_color=BANNER_BG, end_color=BANNER_BG, fill_type="solid")
    cell.font      = FONT_TITLE
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[1].height = 42

    # ── Fila 2: Subtitle con fecha + tipo de búsqueda ────────────────────────
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=num_cols)
    tipo_label = {"bachiller": "Bachiller (Min. Educación)",
                  "satje":     "Procesos Judiciales (SATJE)",
                  "setec":     "Capacitaciones SETEC (Min. del Trabajo)",
                  "completo":  "Verificación completa (Bachiller + SATJE)"}[tipo]
    if incluir_setec and tipo != "setec":
        tipo_label += " + SETEC"
    if incluir_fiscalia:
        tipo_label += " + Fiscalía"
    fecha = datetime.now(_TZ_EC).strftime("%d/%m/%Y %H:%M")
    sub = ws.cell(row=2, column=1, value=f"{tipo_label}  ·  Generado: {fecha}  ·  Powered by JR Automata")
    sub.fill      = PatternFill(start_color=BANNER_BG, end_color=BANNER_BG, fill_type="solid")
    sub.font      = FONT_SUB
    sub.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[2].height = 22

    # ── Fila 3: Stats en una sola línea con contraste fuerte ─────────────────
    if tipo == "completo":
        aptos          = sum(1 for r in resultados if "APTO"        in r.get("semaforo", ""))
        observaciones  = sum(1 for r in resultados if "OBSERVACIÓN" in r.get("semaforo", ""))
        rechazar       = sum(1 for r in resultados if "RECHAZAR"    in r.get("semaforo", ""))
        criticos       = sum(1 for r in resultados if "CRÍTICO"     in r.get("semaforo", ""))
        sin_datos      = sum(1 for r in resultados if "SIN DATOS"   in r.get("semaforo", ""))

        stats_text = (
            f"📊 TOTAL: {len(resultados)}     "
            f"🟢 {aptos} APTOS     "
            f"🟡 {observaciones} OBSERVACIÓN     "
            f"🔴 {rechazar} RECHAZAR     "
            f"🚨 {criticos} CRÍTICOS     "
            f"⚪ {sin_datos} SIN DATOS"
        )
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=num_cols)
        st = ws.cell(row=3, column=1, value=stats_text)
        # Navy oscuro para mejor contraste
        st.fill      = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
        # Segoe UI Emoji renderiza los emojis 🟢🟡🔴⚪🚨 a color real
        st.font      = Font(name="Segoe UI Emoji", size=13, bold=True, color="FFFFFF")
        st.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[3].height = 34
        header_row_num = 4
    else:
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=num_cols)
        st = ws.cell(row=3, column=1, value=f"📊 TOTAL CONSULTADOS: {len(resultados)}")
        st.fill      = PatternFill(start_color=NAVY_LIGHT, end_color=NAVY_LIGHT, fill_type="solid")
        st.font      = FONT_STAT
        st.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[3].height = 32
        header_row_num = 4

    # ── Headers ──────────────────────────────────────────────────────────────
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=header_row_num, column=i, value=h)
        c.fill      = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
        c.font      = FONT_HEADER
        c.alignment = ALIGN_CENTER
        c.border    = _BORDER
    ws.row_dimensions[header_row_num].height = 38

    # ── Anchos de columna ────────────────────────────────────────────────────
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Filas de datos ───────────────────────────────────────────────────────
    def _lbl_bach(estado: str) -> str:
        return {
            "ENCONTRADO":     "✓ Registrado",
            "NO_ENCONTRADO":  "— No registrado",
            "ERROR":          "⚠ Error Min. Educación",
        }.get(estado, estado)

    def _lbl_satje(estado: str) -> str:
        return {
            "SIN_PROCESOS":   "✓ Sin procesos",
            "TIENE_PROCESOS": "⚠ TIENE PROCESOS",
            "ERROR":          "⚠ Error SATJE · F. Judicial",
        }.get(estado, estado)

    for idx, r in enumerate(resultados, start=header_row_num + 1):
        b  = r.get("bachiller", {}) or {}
        s  = r.get("satje", {}) or {}
        st = r.get("setec", {}) or {}
        fi = r.get("fiscalia", {}) or {}
        # Fallback de nombre (defensa adicional al chain del processor):
        #   1) input del Excel → 2) bachiller → 3) SATJE → 4) SETEC
        # En lote, el processor ya aplica este chain y guarda r["nombre"], así
        # que esto sólo entra si por alguna razón vino vacío (cache muy antiguo,
        # registros sin SETEC, etc.).
        nombre = (
            r.get("nombre", "") or
            b.get("nombre", "") or
            s.get("nombre", "") or
            st.get("nombre", "") or
            ""
        )

        # Helpers SETEC para columnas extra
        st_estado_lbl = {
            "TIENE_CERTIFICADOS": "✓ TIENE CERTIFICADOS",
            "SIN_CERTIFICADOS":   "— Sin registros",
            "ERROR":              "⚠ Error SETEC · Min. Trabajo",
            "":                   "",
        }.get(st.get("estado", ""), st.get("estado", ""))
        st_total  = st.get("total", "") if st else ""
        st_cursos = " | ".join(st.get("cursos", []) or []) if st else (st.get("detalle", "") if st else "")

        if tipo == "bachiller":
            fila = [
                r.get("cedula", ""), nombre,
                _lbl_bach(b.get("estado", "")),
                b.get("titulo", ""),
                b.get("especialidad", ""),
                b.get("institucion", ""),
                (b.get("fecha_grado") or "")[:4],
                b.get("detalle", "") if b.get("estado") != "ERROR" else b.get("detalle", "Error de conexión con el servidor del Ministerio de Educación"),
            ]
        elif tipo == "satje":
            fila = [
                r.get("cedula", ""), nombre,
                _lbl_satje(s.get("estado", "")),
                s.get("total_demandado", ""),
                s.get("total_actor", ""),
                s.get("detalle", "") if s.get("estado") != "ERROR" else s.get("detalle", "Error de conexión con el servidor SATJE"),
            ]
        elif tipo == "setec":
            fila = [
                r.get("cedula", ""), nombre,
                st_estado_lbl,
                st_total,
                st_cursos,
            ]
        else:
            sem = r.get("semaforo", "")
            # FIX: armar titulo+especialidad sin .strip(" ()") que come el ) final
            b_titulo = (b.get('titulo') or '').strip()
            b_esp    = (b.get('especialidad') or '').strip()
            if b_titulo and b_esp:
                bach_cell = f"{b_titulo} ({b_esp})"
            elif b_titulo:
                bach_cell = b_titulo
            else:
                bach_cell = b.get("estado", "")

            # Enriquecer "Delitos / Detalle": detalle viejo + delitos como ACTOR
            from src.delitos_clasificacion import resumir_delitos
            detalle_demandado = s.get("detalle", "")
            causas_actor = s.get("causas_actor") or []
            resumen_actor = resumir_delitos(causas_actor)
            if resumen_actor:
                detalle_completo = f"{detalle_demandado} · 🟡 Actor: {resumen_actor}".strip(" ·")
            else:
                detalle_completo = detalle_demandado

            fila = [
                r.get("cedula", ""), nombre, sem,
                bach_cell if b.get("estado") != "ERROR" else "⚠ Error Min. Educación",
                b.get("institucion", ""),
                (b.get("fecha_grado") or "")[:4],
                _lbl_satje(s.get("estado", "")),
                s.get("total_demandado", ""),
                s.get("total_actor", ""),
                detalle_completo,
            ]

        # Helpers Fiscalía para columnas extra
        fi_estado_lbl = {
            "SOSPECHOSO":        "⚠ SOSPECHOSO",
            "DENUNCIANTE":       "ℹ DENUNCIANTE",
            "SIN_ANTECEDENTES":  "✓ Sin antecedentes",
            "ERROR":             "⚠ Error Fiscalía · SIAF",
            "":                  "",
        }.get(fi.get("estado", ""), fi.get("estado", ""))
        fi_total  = fi.get("total_noticias", "") if fi else ""
        fi_delitos = " | ".join(fi.get("delitos", []) or []) if fi else ""

        # Anexar columnas SETEC al final si corresponde
        if incluir_setec and tipo != "setec":
            fila += [st_estado_lbl, st_total, st_cursos]

        # Anexar columnas Fiscalía al final si corresponde
        if incluir_fiscalia:
            fila += [fi_estado_lbl, fi_total, fi_delitos]

        # Banda alternada
        es_par = (idx - header_row_num) % 2 == 0
        bg_color = ROW_ALT_BG if es_par else "FFFFFF"

        for i, val in enumerate(fila, start=1):
            c = ws.cell(row=idx, column=i, value=val)
            c.fill      = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
            c.font      = FONT_CEDULA if i == 1 else FONT_BODY
            c.alignment = ALIGN_LEFT if i in (2, 4, 5, 6, 10) else ALIGN_CENTER
            c.border    = _BORDER

        # Color del nivel de riesgo
        if tipo == "completo":
            sem = r.get("semaforo", "")
            sem_cell = ws.cell(row=idx, column=3)
            if "APTO" in sem:
                sem_cell.fill = PatternFill(start_color=VERDE_BG,    end_color=VERDE_BG,    fill_type="solid")
                sem_cell.font = FONT_BODY_B
            elif "OBSERVACIÓN" in sem:
                sem_cell.fill = PatternFill(start_color=AMARILLO_BG, end_color=AMARILLO_BG, fill_type="solid")
                sem_cell.font = FONT_BODY_B
            elif "CRÍTICO" in sem:
                # Rojo intenso oscuro + texto blanco
                sem_cell.fill = PatternFill(start_color=CRITICO_BG, end_color=CRITICO_BG, fill_type="solid")
                sem_cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
            elif "RECHAZAR" in sem:
                sem_cell.fill = PatternFill(start_color=ROJO_BG,     end_color=ROJO_BG,     fill_type="solid")
                sem_cell.font = FONT_BODY_B
            elif "SIN DATOS" in sem:
                sem_cell.fill = PatternFill(start_color=GRIS_BG,     end_color=GRIS_BG,     fill_type="solid")
                sem_cell.font = FONT_BODY_B
            sem_cell.alignment = ALIGN_CENTER

        # Resaltar SATJE con procesos en rojo
        if tipo in ("satje", "completo"):
            col_estado_satje = 3 if tipo == "satje" else 7
            estado_cell = ws.cell(row=idx, column=col_estado_satje)
            if "TIENE_PROCESOS" in str(estado_cell.value):
                estado_cell.fill = PatternFill(start_color=ROJO_BG, end_color=ROJO_BG, fill_type="solid")
                estado_cell.font = FONT_BODY_B

        # Colorizar columna "Estado SETEC" según resultado
        col_setec_estado = None
        if tipo == "setec":
            col_setec_estado = 3
        elif incluir_setec:
            # Si hay Fiscalía también, SETEC es 3 cols antes del final
            offset = 3 if not incluir_fiscalia else 6
            col_setec_estado = num_cols - offset + 1
        if col_setec_estado:
            setec_cell = ws.cell(row=idx, column=col_setec_estado)
            valor = str(setec_cell.value or "")
            if "TIENE" in valor:
                setec_cell.fill = PatternFill(start_color=VERDE_BG, end_color=VERDE_BG, fill_type="solid")
                setec_cell.font = FONT_BODY_B
            elif "Error" in valor or "✕" in valor:
                setec_cell.fill = PatternFill(start_color=GRIS_BG, end_color=GRIS_BG, fill_type="solid")
                setec_cell.font = FONT_BODY_B

        # Colorizar columna "Estado Fiscalía" según resultado
        if incluir_fiscalia:
            col_fi = num_cols - 2  # 3 cols Fiscalía al final → estado en antepenúltima
            fi_cell = ws.cell(row=idx, column=col_fi)
            valor_fi = str(fi_cell.value or "")
            if "SOSPECHOSO" in valor_fi:
                fi_cell.fill = PatternFill(start_color=ROJO_BG, end_color=ROJO_BG, fill_type="solid")
                fi_cell.font = FONT_BODY_B
            elif "DENUNCIANTE" in valor_fi:
                fi_cell.fill = PatternFill(start_color=AMARILLO_BG, end_color=AMARILLO_BG, fill_type="solid")
                fi_cell.font = FONT_BODY_B
            elif "Sin antecedentes" in valor_fi:
                fi_cell.fill = PatternFill(start_color=VERDE_BG, end_color=VERDE_BG, fill_type="solid")
                fi_cell.font = FONT_BODY_B
            elif "Error" in valor_fi or "✕" in valor_fi:
                fi_cell.fill = PatternFill(start_color=GRIS_BG, end_color=GRIS_BG, fill_type="solid")
                fi_cell.font = FONT_BODY_B

        ws.row_dimensions[idx].height = 32

    # Congelar filas de header + branding
    ws.freeze_panes = ws.cell(row=header_row_num + 1, column=3)  # también freezea cédula+nombre

    # Auto-filter en la tabla
    last_col_letter = get_column_letter(num_cols)
    ws.auto_filter.ref = f"A{header_row_num}:{last_col_letter}{header_row_num + len(resultados)}"

    # Guardar
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def generar_excel_plantilla() -> bytes:
    """Plantilla con look premium para que el usuario llene cédulas + nombres opcionales."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Cédulas"

    # Banner
    ws.merge_cells("A1:B1")
    c = ws.cell(row=1, column=1, value="⚙️ JR Verifica EC — Plantilla")
    c.fill      = PatternFill(start_color=BANNER_BG, end_color=BANNER_BG, fill_type="solid")
    c.font      = FONT_TITLE
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[1].height = 42

    # Instrucciones
    ws.merge_cells("A2:B2")
    s = ws.cell(row=2, column=1, value="Pega tus cédulas en columna A. El nombre es opcional.")
    s.fill      = PatternFill(start_color=BANNER_BG, end_color=BANNER_BG, fill_type="solid")
    s.font      = FONT_SUB
    s.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[2].height = 22

    # Headers
    headers = ["cedula", "nombre"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=3, column=i, value=h)
        c.fill      = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
        c.font      = FONT_HEADER
        c.alignment = ALIGN_CENTER
        c.border    = _BORDER
    ws.row_dimensions[3].height = 32

    # Ejemplos
    ejemplos = [
        ("0954008272", "Jostin Rendón (ejemplo)"),
        ("1309022935", "Fito (ejemplo)"),
    ]
    for i, (ced, nom) in enumerate(ejemplos, start=4):
        ws.cell(row=i, column=1, value=ced).font = FONT_CEDULA
        ws.cell(row=i, column=2, value=nom).font = FONT_BODY
        for col in (1, 2):
            cc = ws.cell(row=i, column=col)
            cc.border    = _BORDER
            cc.alignment = ALIGN_LEFT

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 40
    ws.freeze_panes = "A4"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
